from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from students.forms import StudentsFrom
from students.models import Students
from tablib import Dataset
from .resources import StudentsResource
from import_export import resources
import xlwt
import csv
import openpyxl
import zipfile36

# Create your views here.


def create(request):
    if request.method == 'POST':
        form = StudentsFrom(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Tambah Data Berhasil')
                return redirect('/')
            except:
                print('')
    else:
        form = Students()
    return render(request, 'view.html', {'form': form})


def view(request):
    students = Students.objects.all()
    data = {
        'title': 'CRUDjango',
        'students': students
    }
    return render(request, 'index.html', data)


def edit(request, id):
    students = Students.objects.get(id=id)
    obj = get_object_or_404(Students, id=id)
    form = StudentsFrom(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Edit Data Berhasil')
        return redirect("/")
    data = {
        'id': obj,
        'title': 'Edit Data',
        'students': students
    }
    return render(request, 'edit.html', data)


def delete(request, id):
    if request.method == 'POST':
        students = Students.objects.get(id=id)
        students.delete()
        messages.success(request, 'Delete Data Berhasil')
        return redirect('/')
    else:
        messages.success(request, 'Delete Data gagal')
        return redirect('/')


def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Students.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Students')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['id', 'f_name', 'f_name', 'email']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = Students.objects.all().values_list('id', 'f_name', 'f_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def export_users_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Students.csv"'

    writer = csv.writer(response)
    writer.writerow(['id', 'f_name', 'f_name', 'email'])

    users = Students.objects.all().values_list('id', 'f_name', 'f_name', 'email')
    for user in users:
        writer.writerow(user)

    return response


def import_excel(request):
    if "GET" == request.method:
        return render(request, 'import.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'import.html', {"excel_data": excel_data})
